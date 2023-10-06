import pandas as pd
import PyPDF2 
import re
from openpyxl import load_workbook

# abrindo arquivo em modo leitura e lendo o binário
pdfArquivo = open(r'C:\Users\ramon\OneDrive\Documentos\GitHub\leitoPDF\RelatóriosUFSCmessetembro.pdf','rb')
#dadosPDF recebe os dados em binário do
dadosPDF = PyPDF2.PdfReader(pdfArquivo)

numeroPaginas = len(dadosPDF.pages)
print('O relatório enviado possui: ' + str(numeroPaginas) + ' páginas\n')


wb = load_workbook('controleManutencao.xlsx')               # abre a planilha excel
wb.active                                                   
abaControle = wb['controle']                                # define a aba a ser trabalhada

for pagina in range(0, 59):                                 # no for é determinado o número de páginas lida do pdf
    numeroPagina = pagina                                   #define o numero da pagina a ser lida
    paginaLida = dadosPDF.pages[pagina]                     #armazena a pagina da itereção em paginaLida no formato
    textoPagina = paginaLida.extract_text()                 #extrai todo o texto da pagina
    padrao_data = r'\d{2}/\d{2}/\d{4}'                      #usa o regex para ser usado como padrão de procura de data
    resultado = re.search(padrao_data, textoPagina)         #procura o texto padrão de data em textoPagina
    data = resultado.group(0)                               # trata a informação de data encontada 


    textoProcurado = r'Equipamentos:(.*?)\n'
    textoEq= re.search(textoProcurado, textoPagina) 
    Elevador =textoEq.group(1).strip()

    abaControle.cell(row=pagina+1, column=4,value=str(data))
    abaControle.cell(row=pagina+1, column=2,value=str(Elevador))
    wb.save('controleManutencao.xlsx')
    




pdfArquivo.close()                                          # fecha o arquivo pdf